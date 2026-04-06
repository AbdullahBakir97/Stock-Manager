"""app/services/export_service.py — CSV export for inventory and transactions."""
from __future__ import annotations
import csv
import os
from datetime import datetime
from typing import Optional

from app.repositories.item_repo import ItemRepository
from app.repositories.transaction_repo import TransactionRepository
from app.models.item import InventoryItem
from app.models.transaction import InventoryTransaction


class ExportService:
    """Service for exporting inventory and transaction data to CSV."""

    def __init__(self):
        self._item_repo = ItemRepository()
        self._txn_repo = TransactionRepository()

    # ── Inventory Export ──────────────────────────────────────────────────────

    def export_inventory_csv(
        self, file_path: str, items: list[InventoryItem] | None = None
    ) -> str:
        """
        Export inventory items to CSV.

        Columns:
        ID, Brand, Name, Color, SKU, Barcode, Price, Stock, Min Stock,
        Status, Category, Model, Part Type, Created, Updated

        Args:
            file_path: Destination CSV path
            items: Optional list of items to export; if None, exports all items

        Returns:
            The file path of the created CSV
        """
        if items is None:
            items = self._item_repo.get_all_items()

        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Write header
            headers = [
                "ID",
                "Brand",
                "Name",
                "Color",
                "SKU",
                "Barcode",
                "Price",
                "Stock",
                "Min Stock",
                "Status",
                "Category",
                "Model",
                "Part Type",
                "Created",
                "Updated",
            ]
            writer.writerow(headers)

            # Write rows
            for item in items:
                # Determine category based on item type
                if item.is_product:
                    category = "Product"
                    model = ""
                    part_type = ""
                else:
                    category = "Matrix"
                    model = item.model_name
                    part_type = item.part_type_name

                # Format brand for matrix items
                brand = item.model_brand if item.model_brand else item.brand

                row = [
                    item.id,
                    brand,
                    item.name,
                    item.color,
                    item.sku or "",
                    item.barcode or "",
                    item.sell_price or "",
                    item.stock,
                    item.min_stock,
                    self._format_status(item.stock, item.min_stock),
                    category,
                    model,
                    part_type,
                    item.created_at,
                    item.updated_at,
                ]
                writer.writerow(row)

        return file_path

    # ── Transaction Export ────────────────────────────────────────────────────

    def export_transactions_csv(
        self,
        file_path: str,
        transactions: list[InventoryTransaction] | None = None,
        limit: int = 5000,
    ) -> str:
        """
        Export transaction history to CSV.

        Columns: ID, Timestamp, Item, Operation, Quantity, Before, After, Note

        Args:
            file_path: Destination CSV path
            transactions: Optional list of transactions; if None, fetches from repo
            limit: Maximum number of transactions to export (default 5000)

        Returns:
            The file path of the created CSV
        """
        if transactions is None:
            transactions = self._txn_repo.get_transactions(limit=limit)

        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Write header
            headers = [
                "ID",
                "Timestamp",
                "Item",
                "Operation",
                "Quantity",
                "Before",
                "After",
                "Note",
            ]
            writer.writerow(headers)

            # Write rows
            for txn in transactions:
                row = [
                    txn.id,
                    txn.timestamp,
                    txn.display_name,
                    txn.operation,
                    txn.quantity,
                    txn.stock_before,
                    txn.stock_after,
                    txn.note or "",
                ]
                writer.writerow(row)

        return file_path

    # ── Low Stock Export ──────────────────────────────────────────────────────

    def export_low_stock_csv(self, file_path: str) -> str:
        """
        Export only items at or below min_stock threshold.

        Columns:
        ID, Brand, Name, Color, SKU, Barcode, Price, Stock, Min Stock,
        Status, Deficit, Category, Model, Part Type, Created, Updated

        Args:
            file_path: Destination CSV path

        Returns:
            The file path of the created CSV
        """
        items = self._item_repo.get_low_stock()

        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Write header
            headers = [
                "ID",
                "Brand",
                "Name",
                "Color",
                "SKU",
                "Barcode",
                "Price",
                "Stock",
                "Min Stock",
                "Status",
                "Deficit",
                "Category",
                "Model",
                "Part Type",
                "Created",
                "Updated",
            ]
            writer.writerow(headers)

            # Write rows
            for item in items:
                # Determine category based on item type
                if item.is_product:
                    category = "Product"
                    model = ""
                    part_type = ""
                else:
                    category = "Matrix"
                    model = item.model_name
                    part_type = item.part_type_name

                # Format brand for matrix items
                brand = item.model_brand if item.model_brand else item.brand

                # Calculate deficit (negative = shortage)
                deficit = item.stock - item.min_stock

                row = [
                    item.id,
                    brand,
                    item.name,
                    item.color,
                    item.sku or "",
                    item.barcode or "",
                    item.sell_price or "",
                    item.stock,
                    item.min_stock,
                    self._format_status(item.stock, item.min_stock),
                    deficit,
                    category,
                    model,
                    part_type,
                    item.created_at,
                    item.updated_at,
                ]
                writer.writerow(row)

        return file_path

    # ── Excel Export ─────────────────────────────────────────────────────────

    def export_inventory_xlsx(
        self, file_path: str, items: list[InventoryItem] | None = None
    ) -> str:
        """Export inventory to a professionally formatted Excel workbook."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise RuntimeError("openpyxl is required for Excel export. "
                               "Install with: pip install openpyxl")

        if items is None:
            items = self._item_repo.get_all_items()

        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A",
                                  fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB",
                               fill_type="solid")

        # Status fills
        status_fills = {
            "OK":       PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
            "LOW":      PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            "CRITICAL": PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid"),
            "OUT":      PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
        }

        headers = ["ID", "Brand", "Name", "Color", "SKU", "Barcode",
                    "Price", "Stock", "Min Stock", "Status", "Category",
                    "Model", "Part Type", "Created", "Updated"]

        # Write header
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Write data
        for r, item in enumerate(items, 2):
            brand = item.model_brand if item.model_brand else item.brand
            category = "Product" if item.is_product else "Matrix"
            model = "" if item.is_product else item.model_name
            pt = "" if item.is_product else item.part_type_name
            status = self._format_status(item.stock, item.min_stock)

            row_data = [item.id, brand, item.name, item.color,
                        item.sku or "", item.barcode or "",
                        item.sell_price or "", item.stock, item.min_stock,
                        status, category, model, pt,
                        item.created_at, item.updated_at]

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = thin_border
                if r % 2 == 0:
                    cell.fill = alt_fill

            # Color the status cell
            status_cell = ws.cell(row=r, column=10)
            if status in status_fills:
                status_cell.fill = status_fills[status]

        # Auto-fit column widths (approximate)
        for col in range(1, len(headers) + 1):
            max_len = len(headers[col - 1])
            for row in range(2, min(len(items) + 2, 52)):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 3, 30)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

        wb.save(file_path)
        return file_path

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _format_status(self, stock: int, min_stock: int) -> str:
        """
        Return human-readable stock status.

        Returns:
            One of: "OK", "LOW", "CRITICAL", or "OUT"
        """
        if stock == 0:
            return "OUT"
        if min_stock <= 0:
            return "OK"
        if stock < min_stock:
            return "CRITICAL"
        return "LOW" if stock <= min_stock else "OK"
