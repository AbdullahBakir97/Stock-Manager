"""
app/services/import_service.py — CSV and Excel import service for inventory.

Handles CSV/XLSX preview, delimiter detection, validation, and bulk imports
with duplicate detection and error handling.
Uses the unified ItemRepository (not legacy ProductRepository).
"""
from __future__ import annotations

import csv
import os
from typing import Optional

from app.repositories.item_repo import ItemRepository
from app.core.logger import get_logger

_log = get_logger(__name__)


class ImportService:
    """Service for importing inventory data from CSV and Excel files."""

    def __init__(self) -> None:
        self._item_repo = ItemRepository()

    # ── CSV Preview ──────────────────────────────────────────────────────────

    def preview_csv(self, file_path: str, max_rows: int = 20) -> dict:
        """Preview the first N rows of a CSV file with auto-detected delimiter."""
        if not os.path.isfile(file_path):
            return {"headers": [], "rows": [], "total_rows": 0, "delimiter": ""}

        delimiter = self._detect_delimiter(file_path)
        headers: list[str] = []
        preview_rows: list[list[str]] = []
        total_rows = 0

        try:
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f, delimiter=delimiter)
                headers = next(reader, [])
                for i, row in enumerate(reader):
                    if i < max_rows:
                        preview_rows.append(row)
                    total_rows += 1
        except (IOError, OSError):
            return {"headers": [], "rows": [], "total_rows": 0, "delimiter": delimiter}

        return {
            "headers": headers,
            "rows": preview_rows,
            "total_rows": total_rows,
            "delimiter": delimiter,
        }

    # ── Excel Preview ────────────────────────────────────────────────────────

    def preview_xlsx(self, file_path: str, max_rows: int = 20) -> dict:
        """Preview the first N rows of an Excel file."""
        if not os.path.isfile(file_path):
            return {"headers": [], "rows": [], "total_rows": 0, "delimiter": ""}
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(c) if c else "" for c in next(rows_iter, [])]
            preview_rows: list[list[str]] = []
            total_rows = 0
            for row in rows_iter:
                total_rows += 1
                if len(preview_rows) < max_rows:
                    preview_rows.append([str(c) if c is not None else "" for c in row])
            wb.close()
            return {"headers": headers, "rows": preview_rows,
                    "total_rows": total_rows, "delimiter": ""}
        except Exception as e:
            _log.error(f"XLSX preview error: {e}")
            return {"headers": [], "rows": [], "total_rows": 0, "delimiter": ""}

    # ── Delimiter Detection ──────────────────────────────────────────────────

    def _detect_delimiter(self, file_path: str) -> str:
        delimiters = [",", ";", "\t"]
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                sample = [f.readline().rstrip("\n"), f.readline().rstrip("\n")]
        except (IOError, OSError):
            return ","
        for d in delimiters:
            counts = [len(line.split(d)) for line in sample if line]
            if counts and len(set(counts)) == 1 and counts[0] > 1:
                return d
        return ","

    # ── Row Validation ───────────────────────────────────────────────────────

    def validate_row(self, row: list, column_map: dict,
                     row_num: int) -> tuple[Optional[dict], Optional[str]]:
        """Validate a single row against the column map."""
        brand_idx = column_map.get("brand")
        name_idx = column_map.get("name")
        if brand_idx is None or name_idx is None:
            return (None, f"Row {row_num}: column_map must include 'brand' and 'name'")

        def col(idx: int, default: str = "") -> str:
            return str(row[idx]).strip() if idx < len(row) else default

        brand = col(brand_idx)
        name = col(name_idx)
        if not brand:
            return (None, f"Row {row_num}: brand is required")
        if not name:
            return (None, f"Row {row_num}: name is required")

        parsed: dict = {"brand": brand, "name": name}

        # Color
        color_idx = column_map.get("color")
        parsed["color"] = col(color_idx) if color_idx is not None else ""

        # Barcode (check duplicates)
        barcode_idx = column_map.get("barcode")
        if barcode_idx is not None:
            bc = col(barcode_idx)
            if bc:
                existing = self._item_repo.get_by_barcode(bc)
                if existing:
                    return (None, f"Row {row_num}: barcode '{bc}' already exists")
                parsed["barcode"] = bc
            else:
                parsed["barcode"] = None
        else:
            parsed["barcode"] = None

        # Numeric fields
        for field, idx_key in [("stock", "stock"), ("min_stock", "min_stock")]:
            idx = column_map.get(idx_key)
            if idx is not None:
                val = col(idx)
                try:
                    parsed[field] = int(val) if val else 0
                except ValueError:
                    return (None, f"Row {row_num}: {field} must be numeric, got '{val}'")
            else:
                parsed[field] = 0

        # Price
        price_idx = column_map.get("price")
        if price_idx is not None:
            pval = col(price_idx)
            try:
                parsed["price"] = float(pval) if pval else None
            except ValueError:
                return (None, f"Row {row_num}: price must be numeric, got '{pval}'")
        else:
            parsed["price"] = None

        return (parsed, None)

    # ── CSV Import ───────────────────────────────────────────────────────────

    def import_products_csv(self, file_path: str, column_map: dict,
                            skip_header: bool = True) -> dict:
        """Import products from a CSV file into inventory_items."""
        return self._import_from_rows(
            self._csv_rows(file_path, skip_header),
            column_map,
            start_row=2 if skip_header else 1,
        )

    # ── Excel Import ─────────────────────────────────────────────────────────

    def import_products_xlsx(self, file_path: str, column_map: dict,
                             skip_header: bool = True) -> dict:
        """Import products from an Excel file into inventory_items."""
        return self._import_from_rows(
            self._xlsx_rows(file_path, skip_header),
            column_map,
            start_row=2 if skip_header else 1,
        )

    # ── Internal import logic ────────────────────────────────────────────────

    def _import_from_rows(self, rows_iter, column_map: dict,
                          start_row: int = 2) -> dict:
        imported = 0; skipped = 0; errors: list[str] = []
        for row_idx, row in enumerate(rows_iter, start=start_row):
            if not row or all(not str(c).strip() for c in row):
                continue
            row_strs = [str(c) if c is not None else "" for c in row]
            parsed, error = self.validate_row(row_strs, column_map, row_idx)
            if error:
                errors.append(error); skipped += 1; continue
            try:
                self._item_repo.add_product(
                    brand=parsed["brand"], name=parsed["name"],
                    color=parsed["color"], stock=parsed["stock"],
                    barcode=parsed["barcode"], min_stock=parsed["min_stock"],
                    sell_price=parsed["price"],
                )
                imported += 1
            except Exception as e:
                errors.append(f"Row {row_idx}: {e}"); skipped += 1
        _log.info(f"Import complete: {imported} imported, {skipped} skipped")
        return {"imported": imported, "skipped": skipped, "errors": errors}

    def _csv_rows(self, file_path: str, skip_header: bool):
        """Yield rows from a CSV file."""
        if not os.path.isfile(file_path):
            return
        delimiter = self._detect_delimiter(file_path)
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f, delimiter=delimiter)
            if skip_header:
                next(reader, None)
            yield from reader

    def _xlsx_rows(self, file_path: str, skip_header: bool):
        """Yield rows from an Excel file."""
        if not os.path.isfile(file_path):
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            if skip_header:
                next(rows_iter, None)
            for row in rows_iter:
                yield list(row)
            wb.close()
        except Exception as e:
            _log.error(f"XLSX import error: {e}")
